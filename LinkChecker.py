# This code checks the links in  series and puts the response in a nearby blank cell (useful for error checking)
# import openpyxl
# import requests
# from openpyxl.styles import PatternFill

# def test_hyperlinks(file_path):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.active
#     rowCount=1

#     for row in(ws.iter_rows(min_row=2, max_col=20, values_only=True)):
#         rowCount+=1
#         cell_value = row[1]
#         if cell_value:
#             try:
#                 response=requests.head(cell_value, allow_redirects=True)
#                 status_code=response.status_code
                             
#                 if 200<=status_code<400:
#                     result="Active"
#                     ws.cell(row=rowCount, column=2).fill = PatternFill(start_color="00FF00", fill_type = "solid")
#                 else:
#                     result=f"Inactive ({status_code})"
#                     ws.cell(row=rowCount, column=2).fill = PatternFill(start_color="FF0000", fill_type = "solid")                  

#             except:
#                     requests.RequestException
#                     result=f"Error"
#                     ws.cell(row=rowCount, column=2).fill = PatternFill(start_color="FFA500", fill_type = "solid")
            
#     wb.save(file_path)

#test_hyperlinks(r"C:\Users\Flip\Desktop\LinkChecker\Copy of CASI Municipalities Masterbook.xlsx")



# This code checks the links in series and changes the background color of the cells
# import openpyxl
# import requests
# from openpyxl.styles import PatternFill

# def test_hyperlinks(file_path):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.active

#     for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=20):
#         for cell in row:
#             hyperlink = cell.hyperlink
#             if hyperlink:
#                 try:
#                     response = requests.head(hyperlink.target, allow_redirects=True, timeout=10)
#                     status_code = response.status_code

#                     print(f"URL: {hyperlink.target}, Status Code: {status_code}")

#                     if 200 <= status_code < 400:
#                         cell.fill = PatternFill(start_color="00FF00", fill_type="solid")
#                     else:
#                         cell.fill = PatternFill(start_color="FF0000", fill_type="solid")

#                 except requests.RequestException:
#                     print(f"Error checking URL: {hyperlink.target}, Exception: {e}")
#                     cell.fill = PatternFill(start_color="FFA500", fill_type="solid")

#     wb.save(file_path)



# test_hyperlinks(r"C:\Users\Flip\Desktop\LinkChecker\Copy of CASI Municipalities Masterbook.xlsx")


# This code checks the links asynchronously and changes the background color of the cells
# Added additional logging to terminal window
import openpyxl
import aiohttp
import asyncio
from openpyxl.styles import PatternFill

#def user_input():
    #path = input("Enter the full directory path to the Excel file we're testing: ").strip()
    #file_name = input("Enter the file name of Excel file we're testing: ").strip()
    #test_hyperlinks(r"{path}+{file_name}")

#This function not currently working
#user_input()

async def check_link(session, cell):
    hyperlink = cell.hyperlink
    try:
        async with session.head(hyperlink.target, allow_redirects=True) as response:
            status_code = response.status

            print(f"URL: {hyperlink.target}, Status Code: {status_code}")

            if 200 <= status_code < 400:
                result = "Active"
                cell.fill = PatternFill(start_color="00FF00", fill_type="solid")
            else:
                result = f"Inactive ({status_code})"
                cell.fill = PatternFill(start_color="FF0000", fill_type="solid")

    except aiohttp.ClientError as e:
        print(f"Error checking URL: {hyperlink.target}, Exception: {e}")
        result = "Error"
        cell.fill = PatternFill(start_color="FFA500", fill_type="solid")

async def test_hyperlinks(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    async with aiohttp.ClientSession() as session:
        tasks = []

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=20):
            for cell in row:
                hyperlink = cell.hyperlink
                if hyperlink:
                    task = check_link(session, cell)
                    tasks.append(task)

        await asyncio.gather(*tasks)

    wb.save(file_path)

asyncio.run(test_hyperlinks(r"C:\Users\Flip\Desktop\LinkChecker\Copy of CASI Municipalities Masterbook.xlsx"))
